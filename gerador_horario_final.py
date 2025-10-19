import pandas as pd
import pdfplumber
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÃO PRINCIPAL ---
# LISTAS DE PROFESSORES ATUALIZADAS E EM ORDEM ALFABÉTICA
NOMES_MASCULINOS = [
    "CONRADO",
    "CRISTIAN",
    "DANILO",
    "DOUGLAS",
    "JAMES",
    "JESSÉ",
    "JOÃO VITOR",
    "JOÃOVITOR",
    "LEANDRO",
    "LUIZ",
    "MARCOS",
    "MARTINI",
]
NOMES_FEMININOS = [
    "ADRIANA",
    "ALINE",
    "AMANDA",
    "CAREM",
    "CHEILA",
    "CLAUDINEIA",
    "DANIELE",
    "EDINEIA",
    "EDSSEIA",
    "FERNANDA",
    "FRANCIELE",
    "JAQUELINE",
    "KAROLINE",
    "KELYÇA",
    "LESLIE",
    "LUANA",
    "MARILZA",
    "NEIRE",
    "PATRÍCIA",
    "ROSELI",
    "SILVANA",
    "SIMONE",
    "SOLANGE",
    "SORAIA",
    "VERA",
    "VITÓRIA",
]

CORES_FRIAS = [
    "B0E0E6",
    "ADD8E6",
    "87CEFA",
    "B0C4DE",
    "778899",
    "AFEEEE",
    "40E0D0",
    "48D1CC",
    "00CED1",
    "5F9EA0",
    "66CDAA",
    "8FBC8F",
    "98FB98",
    "90EE90",
    "3CB371",
    "2E8B57",
    "BDB76B",
    "B4D3B2",
]
CORES_QUENTES = [
    "FFDAB9",
    "FFE4B5",
    "F0E68C",
    "FFB6C1",
    "FFA07A",
    "FFC0CB",
    "FFDEAD",
    "F5DEB3",
    "DAA520",
    "FF69B4",
    "FF6347",
    "FF4500",
    "F4A460",
    "D2B48C",
    "BC8F8F",
    "E9967A",
    "CD5C5C",
    "DB7093",
    "C71585",
    "FFA500",
    "FF8C00",
    "FFD700",
]

# --- MAPEAMENTO ATUALIZADO PARA O NOVO PDF ---
MAPEAMENTO_LOCAL = {
    (1, 0): "Biblioteca",
    (1, 1): "Lab. Ciências",
    (2, 0): "Lab. Informática",
    (2, 1): "Lab. Robótica - Multiuso",
    (3, 0): "Sala 1",
    (3, 1): "Sala 10",
    (3, 2): "Sala 11",
    (4, 0): "Sala 12",
    (4, 1): "Sala 2",
    (5, 0): "Sala 3",
    (5, 1): "Sala 4",
    (5, 2): "Sala 5",
    (6, 0): "Sala 6",
    (6, 1): "Sala 7",
    (7, 0): "Sala 9",
    (7, 1): "Sala 8",
}

# --- LISTA DE LOCAIS ATUALIZADA E NA ORDEM DESEJADA ---
LISTA_FIXA_LOCAIS = [
    "Sala 1",
    "Sala 2",
    "Sala 3",
    "Sala 4",
    "Sala 5",
    "Sala 6",
    "Sala 7",
    "Sala 8",
    "Sala 9",
    "Sala 10",
    "Sala 11",
    "Sala 12",
    "Lab. Robótica - Multiuso",
    "Lab. Informática",
    "Lab. Ciências",
    "Biblioteca",
]

CAMINHO_DO_SCRIPT = (
    os.path.dirname(os.path.abspath(__file__))
    if "__file__" in locals()
    else os.getcwd()
)
PDF_HORARIOS = os.path.join(CAMINHO_DO_SCRIPT, "horarios.pdf")
NOME_PLANILHA_SAIDA = os.path.join(CAMINHO_DO_SCRIPT, "Horario_ATUALIZADO.xlsx")

# --- FUNÇÕES (O restante do seu código original) ---


def limpar_texto(texto):
    return str(texto).replace("\n", " ").strip() if texto is not None else ""


def criar_mapa_de_cores():
    mapa = {}

    # Mapeia cores para nomes masculinos
    nomes_masculinos_unicos = sorted(list(set(NOMES_MASCULINOS)))
    for i, nome in enumerate(nomes_masculinos_unicos):
        mapa[nome.upper()] = CORES_FRIAS[i % len(CORES_FRIAS)]

    # Mapeia cores para nomes femininos
    nomes_femininos_unicos = sorted(list(set(NOMES_FEMININOS)))
    for i, nome in enumerate(nomes_femininos_unicos):
        mapa[nome.upper()] = CORES_QUENTES[i % len(CORES_QUENTES)]

    return mapa


def extrair_dados_pdf_recursos(nome_pdf):
    dados_extraidos = []
    with pdfplumber.open(nome_pdf) as pdf:
        for num_pagina, pagina in enumerate(pdf.pages, 1):
            tabelas = pagina.extract_tables()
            if not tabelas:
                continue
            for i, tabela in enumerate(tabelas):
                local_atual = MAPEAMENTO_LOCAL.get((num_pagina, i))
                if not local_atual:
                    continue
                dia_da_semana_atual = None
                for linha in tabela[1:]:  # Pula cabeçalho
                    if not linha or not any(linha):
                        continue
                    primeira_celula = limpar_texto(linha[0])

                    # Regex para encontrar dias da semana de forma mais robusta
                    match_dia = re.match(
                        r"^(Segunda|Terça|Quarta|Quinta|Sexta)",
                        primeira_celula,
                        re.IGNORECASE,
                    )
                    if match_dia:
                        dia_da_semana_atual = match_dia.group(0).strip() + "-feira"
                        # Caso especial para "Quarta" que às vezes não tem hifen
                        if dia_da_semana_atual == "Quarta-feira":
                            dia_da_semana_atual = "Quarta-feira"

                    # Verifica se é uma linha de dados válida
                    if (
                        dia_da_semana_atual
                        and linha
                        and primeira_celula
                        and re.match(r"^\d{2}:\d{2}", primeira_celula)
                    ):
                        try:
                            dados_aula = {
                                "Local": local_atual,
                                "Dia": dia_da_semana_atual,
                                "Início": limpar_texto(linha[0]),
                                "Aulas": limpar_texto(linha[2]),
                                "Turma": limpar_texto(linha[3]),
                                "Disciplina": limpar_texto(linha[4]),
                                "Professor": limpar_texto(linha[5]),
                            }
                            if dados_aula["Início"] and dados_aula["Turma"]:
                                dados_extraidos.append(dados_aula)
                        except (IndexError, TypeError):
                            continue
    return dados_extraidos


def expandir_aulas_geminadas(df):
    horarios_inicio = [
        "07:35",
        "08:25",
        "09:30",
        "10:20",
        "11:10",
        "13:00",
        "13:50",
        "14:55",
        "15:45",
    ]
    novas_linhas = []
    for _, row in df.iterrows():
        inicio_limpo = (
            re.sub(r"[^0-9:]", "", row["Início"])
            if isinstance(row["Início"], str)
            else ""
        )
        if not inicio_limpo:
            continue

        row["Início"] = inicio_limpo  # Normaliza o horário na linha original
        novas_linhas.append(row)

        # Lógica para aulas duplas, triplas etc.
        try:
            num_aulas = int(row["Aulas"])
        except (ValueError, TypeError):
            num_aulas = 1

        if num_aulas > 1:
            try:
                idx = horarios_inicio.index(inicio_limpo)
                for i in range(1, num_aulas):
                    if idx + i < len(horarios_inicio):
                        nova_linha = row.copy()
                        nova_linha["Início"] = horarios_inicio[idx + i]
                        novas_linhas.append(nova_linha)
            except ValueError:
                continue
    return pd.DataFrame(novas_linhas)


def formatar_e_colorir_planilha(nome_arquivo, df_professores, mapa_de_cores):
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    fills = {
        cor: PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        for cor in set(mapa_de_cores.values())
    }
    font_padrao, font_bold = Font(name="Calibri", size=10), Font(
        name="Calibri", size=10, bold=True
    )
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = borda

    for cell in sheet[1]:
        cell.font = font_bold

    for r_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), 2):
        for c_idx, cell in enumerate(row_cells, 1):
            cell.alignment = alinhamento
            cell.font = font_bold if c_idx <= 2 else font_padrao
            if c_idx > 2:
                try:
                    # Normaliza o nome do professor encontrado na planilha para busca
                    prof_nome_sujo = df_professores.iloc[r_idx - 2, c_idx - 3]
                    if pd.notna(prof_nome_sujo):
                        prof_nome_limpo = re.sub(
                            r"[^A-Z\s]", "", str(prof_nome_sujo).upper()
                        ).strip()
                        cor = mapa_de_cores.get(prof_nome_limpo)
                        if cor:
                            cell.fill = fills[cor]
                except IndexError:
                    continue

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 8
    for i in range(3, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 15

    workbook.save(nome_arquivo)


def mapear_horario_para_aula(inicio):
    inicio_limpo = re.sub(r"[^0-9:]", "", inicio) if isinstance(inicio, str) else ""
    return {
        "07:35": "1ª aula",
        "08:25": "2ª aula",
        "09:30": "3ª aula",
        "10:20": "4ª aula",
        "11:10": "5ª aula",
        "13:00": "6ª aula",
        "13:50": "7ª aula",
        "14:55": "8ª aula",
        "15:45": "9ª aula",
    }.get(inicio_limpo)


def abreviar(texto, tamanho=5):
    if pd.isna(texto):
        return ""
    # Remove acentos e caracteres especiais antes de abreviar
    texto_limpo = str(texto).upper()
    return texto_limpo[:tamanho]


def main():
    if not os.path.exists(PDF_HORARIOS):
        print(
            f"\n--- ERRO ---\nO arquivo '{os.path.basename(PDF_HORARIOS)}' não foi encontrado."
        )
        print(
            "Por favor, renomeie o novo PDF para 'horarios.pdf' e coloque-o na mesma pasta do script."
        )
        return

    print("Iniciando a leitura e processamento do PDF...")
    df_bruto = pd.DataFrame(extrair_dados_pdf_recursos(PDF_HORARIOS))
    if df_bruto.empty:
        print(
            "\nNenhum dado válido foi extraído. Verifique o `MAPEAMENTO_LOCAL` e a estrutura do PDF."
        )
        return

    df_horario = expandir_aulas_geminadas(df_bruto)

    # Normaliza nomes de professores para remover variações de acento/capitalização
    df_horario["Professor"] = df_horario["Professor"].str.upper().str.strip()
    # Corrige possíveis erros de digitação comuns
    df_horario["Professor"] = df_horario["Professor"].replace(
        {"MARITZA": "MARILZA", "MANIZA": "MARILZA", "CAREM": "KAREM"}
    )

    mapa_de_cores = criar_mapa_de_cores()

    prof_extraidos = set(df_horario["Professor"].dropna().unique())
    prof_mapeados = set(mapa_de_cores.keys())
    prof_faltantes = prof_extraidos - prof_mapeados

    if prof_faltantes:
        print("\n--- ATENÇÃO: Professores não encontrados no código! ---")
        print("A planilha foi gerada, mas estes professores não serão coloridos:")
        for nome in sorted(list(prof_faltantes)):
            if nome:
                print(f"- {nome}")
        print("----------------------------------------------------------")

    df_horario["Nº aula"] = df_horario["Início"].apply(mapear_horario_para_aula)
    df_horario.dropna(subset=["Nº aula"], inplace=True)

    df_horario["Valor Célula"] = df_horario.apply(
        lambda row: f"{row['Turma']}\n{abreviar(row['Disciplina'])}-{abreviar(row['Professor'])}",
        axis=1,
    )

    print("\nMontando a matriz de horários...")
    df_matriz_display = df_horario.pivot_table(
        index=["Dia", "Nº aula"],
        columns="Local",
        values="Valor Célula",
        aggfunc="first",
    )
    df_matriz_professores = df_horario.pivot_table(
        index=["Dia", "Nº aula"], columns="Local", values="Professor", aggfunc="first"
    )

    ordem_aulas = [f"{i}ª aula" for i in range(1, 10)]
    ordem_dias = [
        "Segunda-feira",
        "Terça-feira",
        "Quarta-feira",
        "Quinta-feira",
        "Sexta-feira",
    ]
    idx = pd.MultiIndex.from_product(
        [ordem_dias, ordem_aulas], names=["Dia", "Nº aula"]
    )

    # Adiciona colunas faltantes e reordena
    for local in LISTA_FIXA_LOCAIS:
        if local not in df_matriz_display.columns:
            df_matriz_display[local] = pd.NA
            df_matriz_professores[local] = pd.NA

    df_matriz_display = df_matriz_display.reindex(idx)[LISTA_FIXA_LOCAIS].reset_index()
    df_matriz_professores = df_matriz_professores.reindex(idx)[LISTA_FIXA_LOCAIS]

    print(f"Salvando a planilha base em '{os.path.basename(NOME_PLANILHA_SAIDA)}'...")
    df_matriz_display.to_excel(
        NOME_PLANILHA_SAIDA, sheet_name="Horário Colorido", index=False
    )
    formatar_e_colorir_planilha(
        NOME_PLANILHA_SAIDA, df_matriz_professores, mapa_de_cores
    )

    print("\n--- ✅ SUCESSO! ---")
    print(
        f"A planilha '{os.path.basename(NOME_PLANILHA_SAIDA)}' foi criada com sucesso."
    )


if __name__ == "__main__":
    main()
